"""Utilities for processing Slocum glider files the Kiwi way"""
import logging
import numpy as np
from openpyxl import load_workbook


_log = logging.getLogger(__name__)


def collect_excelsheet_metadata(
        excelsheet: str, ID: int = 1
) -> dict[str, list[str | int | float]]:
    """Extract Excel sheet metadata for given deployment ID.

    Note
    ----
    First row of Excel sheet must contain descriptive column headers.
    First column of Excel sheet must contain deployment ID numbers.

    Parameters
    ----------
    excelsheet : str
        Path to Excel sheet with metadata
    ID : int
        Deployment number/ID to focus on

    Returns
    -------
    dict[str, list[str | int | float]]
        Excel sheet metadata from specified deployment

    """
    _log.info(f'Getting metadata from {excelsheet}')

    # load the Excel sheet
    worksheet = load_workbook(excelsheet).active
    # get header names (must be present in first row)
    headers = [
        worksheet.cell(row=1, column=idx).value
        for idx in range(2, worksheet.max_column+1)
    ]
    # get deployment numbers (must be present in first column)
    deployments = [
        worksheet.cell(row=idx, column=1).value
        for idx in range(2, worksheet.max_row+1)
    ]

    # return metadata from input deployment ID
    return {
        headers[idx-2]: worksheet.cell(
            row=deployments.index(ID)+2, column=idx
        ).value
        for idx in range(2, worksheet.max_column+1)
    }


def dd2dm(decimal_degrees: float) -> tuple[float, float]:
    """Convert decimal degrees to degree minute notation

    Note
    ----
    adapted from MATLAB's degrees2dm function

    Paramters
    ---------
    decimal_degrees : float
        Coordinate, either longitude or latitude, in decimal degrees

    Returns
    -------
    float
        Degree portion of input coordinate
    float
        Minute portion of input coordinate

    """
    degrees = np.fix(decimal_degrees)
    minutes = 60*np.remainder(np.abs(decimal_degrees), 1)
    return degrees, minutes


def dm2dd(degrees: float, minutes: float) -> float:
    """Convert degree minute to decimal degrees notation

    Note
    ----
    adapted from MATLAB's dm2degrees function

    Paramters
    ---------
    degrees : float
        Degree portion of input coordinate, either latitude or longitude
    minutes : float
        Minute portion of input coordinate, either latitude or longitude

    Returns
    -------
    float
        Coordinate in decimal degrees

    """
    decimal_degrees = np.abs(degrees) + np.abs(minutes)/60
    decimal_degrees = np.where(degrees < 0 or minutes < 0, decimal_degrees*-1,
                               decimal_degrees)
    return decimal_degrees


def first_nonnan(numpy_array: np.array) -> float:
    """Get the first non-nan value in a numpy array.

    Parameters
    ----------
    numpy_array : numpy array
        Numpy array to search

    Returns
    -------
    float
        First non-nan value in array

    """
    return numpy_array[np.isfinite(numpy_array)][0]


def last_nonnan(numpy_array: np.array) -> float:
    """Get the last non-nan value in a numpy array.

    Parameters
    ----------
    numpy_array : numpy array
        Numpy array to search

    Returns
    -------
    float
        Last non-nan value in array

    """
    return numpy_array[np.isfinite(numpy_array)][-1]
