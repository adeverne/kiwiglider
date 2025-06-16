"""
Classes to process Slocum glider files the (basic) Kiwi way
"""
from os.path import join as join_path
from os.path import basename, exists
from os import makedirs, listdir
from pathlib import Path
from tempfile import NamedTemporaryFile
from itertools import groupby
import yaml
import logging
from typing import Any
from openpyxl import load_workbook
from pyglider import slocum
from pyglider.ncprocess import extract_timeseries_profiles
from ioos_qc.config import Config
from ioos_qc.streams import XarrayStream
from ioos_qc.results import collect_results, ContextResult, CallResult
from ioos_qc.qartod import aggregate
from ioos_qc.stores import PandasStore, column_from_collected_result
from compliance_checker.runner import ComplianceChecker, CheckSuite
import xarray as xr
import numpy as np
from datetime import datetime
import pygmt
from utm import from_latlon
from inspect import getmodule
from kiwiglider.utils import dd2dm, temporary_cpt, first_nonnan, last_nonnan


_log = logging.getLogger(__name__)


def collect_excelsheet_metadata(
        excelsheet: str, ID: int = 1
) -> dict[str, list[str | int | float]]:
    """Extract Excel sheet metadata for given deployment ID.

    Note
    ----
    First row of Excel sheet must contain descriptive column headers.
    First column of Excel sheet must contain deployment ID numbers.

    Parameters
    ----------
    excelsheet : str
        Path to Excel sheet with metadata
    ID : int
        Deployment number/ID to focus on

    Returns
    -------
    dict[str, list[str | int | float]]
        Excel sheet metadata from specified deployment

    """
    _log.info(f'Getting metadata from {excelsheet}')

    # load the Excel sheet
    worksheet = load_workbook(excelsheet).active
    # get header names (must be present in first row)
    headers = [
        worksheet.cell(row=1, column=idx).value
        for idx in range(2, worksheet.max_column+1)
    ]
    # get deployment numbers (must be present in first column)
    deployments = [
        worksheet.cell(row=idx, column=1).value
        for idx in range(2, worksheet.max_row+1)
    ]

    # return metadata from input deployment ID
    return {
        headers[idx-2]: worksheet.cell(
            row=deployments.index(ID)+2, column=idx
        ).value
        for idx in range(2, worksheet.max_column+1)
    }


class DeploymentYAML():
    """Contains deployment-specific metadata.

    Parameters
    ----------
    ID : int, default = 1
        Deployment number/ID to focus on

    Attributes
    ----------
    ID : int
        Deployment number/ID to focus on
    excelsheet : str
        Path to Excel sheet with metadata
    excel_metadata : dict[str, list[str | int | float]]
        Excel sheet metadata from specified deployment
    metadata : dict
        Global deployment metadata
    glider_devices : dict
        Installed instrument metadata
    netcdf_variables : dict
        NetCDF variable name mapping and metadata
    profile_variables : dict
        Profile-averaged variable metadata
    qartod_tests : dict
        QARTOD test parameters
    yaml : dict
        All metadata to go in YAML file
    outname : str
        Path to output YAML file

    """

    def __init__(self, ID: int = 1):
        self.ID = ID

    def add_excel_metadata(self, excelsheet: str) -> None:
        """Add Excel sheet metadata for given deployment ID.

        Note
        ----
        First row of Excel sheet must contain descriptive column headers.
        First column of Excel sheet must contain deployment ID numbers.

        Parameters
        ----------
        excelsheet : str
            Path to Excel sheet with metadata

        """

        # assign input Excel sheet
        self.excelsheet = excelsheet

        # get metadata from input deployment ID
        self.excel_metadata = collect_excelsheet_metadata(
            self.excelsheet, self.ID
        )

    def add_metadata(self, metadata: dict[str, Any] = None) -> None:
        """Add global deployment metadata.

        Parameters
        ----------
        metadata : dict[str, Any], optional
            Global attributes to overwrite default
            in form {'attribute':value}

        """
        _log.info('Adding global metadata to deployment YAML')

        # make sure already have Excel sheet metadata
        self._check_for_excel_meta()

        # initialize with static metadata
        self.metadata = {
            'Conventions': 'CF-1.11',
            'Metadata_Conventions': 'CF-1.11, Unidata Dataset Discovery v1.0',
            'contributor_role_vocabulary': 'http://vocab.nerc.ac.uk' +
            '/search_nvs/W08/',
            'comment': '" "',
            'creator_url': '" "',
            'format_version': 'IOOS_Glider_NetCDF_v2.0.nc',
            'keywords':
                'Water-based Platforms > Uncrewed Vehicles > Subsurface > ' +
                'Seaglider, Oceans > Marine Sediments > Turbidity, ' +
                'Oceans > Ocean Chemistry > Oxygen, ' +
                'Oceans > Ocean Circulation > Turbulence, ' +
                'Oceans > Ocean Pressure > Water Pressure, ' +
                'Oceans > Ocean Temperature > Water Temperature, ' +
                'Oceans > Salinity/Density > Conductivity, ' +
                'Oceans > Salinity/Density > Density, ' +
                'Oceans > Salinity/Density > Salinity',
            'keywords_vocabulary': 'GCMD Science Keywords',
            'license': 'This data may be redistributed and used without ' +
            'restriction',
            'metadata_link': '" "',
            'processing_level': 'Data are provided as-is',
            'publisher_url': '" "',
            'references': '" "',
            'source': 'Observational data from a profiling glider',
            'standard_name_vocabulary': 'Standard Name Table (v85, 21 May ' +
            '2024)',
            'summary': 'This dataset contains physical oceanographic ' +
            'measurements of temperature, conductivity, salinity, density ' +
            'and estimates of depth-average currents.',
        }

        # add from Excel sheet
        acknowledge = ('This work supported by funding from ' +
                       self.excel_metadata['funding'])
        self._add_meta('acknowledgement', acknowledge)
        pi = self.excel_metadata['principal_investigator']
        dm = self.excel_metadata['data_manager']
        pilot = self.excel_metadata['pilot']
        contributor_name = ','.join([pi, dm, pilot])
        self._add_meta('contributor_name', contributor_name)
        contributor_role = ','.join(
            ['Principal Investigator']*int(pi.count(',')+1) +
            ['Data Manager']*int(dm.count(',')+1) +
            ['Operator']*int(pilot.count(',')+1)
        )
        self._add_meta('contributor_role', contributor_role)
        self._add_meta('creator_email',
                       self.excel_metadata['data_manager_email'])
        self._add_meta('creator_name', self.excel_metadata['data_manager'])
        self._add_meta('deployment_name', 'GLD{:04d}'.format(self.ID))
        self._add_meta('deployment_start',
                       self.excel_metadata['deploy_date'].strftime('%Y-%m-%d'))
        # note: will be overwritten in nc by pyglider
        self._add_meta('deployment_end',
                       self.excel_metadata['end_date'].strftime('%Y-%m-%d'))
        # note: will be overwritten in nc by pyglider
        self._add_meta('glider_name', self.excel_metadata['platform_id'])
        self._add_meta('glider_serial',
                       f'{self.excel_metadata['platform_sn']}')
        self._add_meta('glider_model', self.excel_metadata['glidertype'])
        self._add_meta('glider_pump', f'{self.excel_metadata['pump_type']}m')
        self._add_meta('institution', self.excel_metadata['owner'])
        backwards_email = self.excel_metadata['data_manager_email']
        backwards_email = backwards_email.split('@')[-1]
        backwards_email = backwards_email.split('.')[-1::-1]
        backwards_email = '.'.join(backwards_email)
        self._add_meta('naming_authority', backwards_email)
        self._add_meta('platform_type', self.excel_metadata['platform_type'])
        self._add_meta('project', self.excel_metadata['project_name'])
        self._add_meta('publisher_email',
                       self.excel_metadata['data_manager_email'])
        self._add_meta('publisher_name', self.excel_metadata['data_manager'])
        self._add_meta('sea_name', self.excel_metadata['sea'])
        self._add_meta('wmo_id', f'{self.excel_metadata['wmo_id']}')

        # add from user input (overwrite above as necessary)
        if metadata is not None:
            for name, value in metadata.items():
                self._add_meta(name, value)

    def add_glider_devices(
            self,
            glider_devices: dict[str, dict[str, Any]] = None
    ) -> None:
        """Add installed instrument metadata.

        Parameters
        ----------
        glider_devices : dict[str, dict[str, Any]], optional
            Glider device metadata to overwrite default
            in form {'device_name': {'attribute': value}}

        """
        _log.info('Adding glider device metadata to deployment YAML')

        # make sure already have Excel sheet metadata
        self._check_for_excel_meta()

        # initialize common instruments
        self.glider_devices = {
            'pressure': {
                'make': 'Micron',
                'model': 'Pressure',
                'serial': f'{self.excel_metadata['pres_sn']}'
            },
            'ctd': {
                'make': 'Seabird',
                'model': self.excel_metadata['ctd_type'],
                'serial': f'{self.excel_metadata['ctd_sn']}',
                'long_name': 'Seabird SlocumCTD',
                'make_model': 'Seabird SlocumCTD',
                'factory_calibrated': '" "',
                'calibration_date':
                self.excel_metadata['ctd_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            }
        }

        # add based on devices present in Excel worksheet metadata
        if self.excel_metadata['wetlabs_installed']:
            self._add_glider_device('optics', {
                'make': 'Wetlabs',
                'model': self.excel_metadata['wetlabs_type'],
                'serial': f'{self.excel_metadata['wetlabs_sn']}',
                'factory_calibrated':
                self.excel_metadata['wetlabs_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['wetlabs_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })
        if self.excel_metadata['oxy_installed']:
            self._add_glider_device('oxygen', {
                'make': 'AADI',
                'model': self.excel_metadata['oxy_type'],
                'serial': f'{self.excel_metadata['oxy_sn']}',
                'factory_calibrated':
                self.excel_metadata['oxy_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['oxy_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })
        if self.excel_metadata['par_installed']:
            self._add_glider_device('par', {
                'make': 'Biospherical',
                'model': self.excel_metadata['par_type'],
                'serial': f'{self.excel_metadata['par_sn']}',
                'factory_calibrated':
                self.excel_metadata['par_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['par_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })
        if self.excel_metadata['bb3_installed']:
            self._add_glider_device('optics2', {
                'make': 'SeaBird',
                'model': self.excel_metadata['bb3_type'],
                'serial': f'{self.excel_metadata['bb3_sn']}',
                'factory_calibrated':
                self.excel_metadata['bb3_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['bb3_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })
        if self.excel_metadata['lisst_installed']:
            self._add_glider_device('lisst', {
                'make': 'Sequoia',
                'model': self.excel_metadata['lisst_type'],
                'serial': f'{self.excel_metadata['lisst_sn']}',
                'factory_calibrated':
                self.excel_metadata['lisst_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['lisst_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })
        if self.excel_metadata['microrider_installed']:
            self._add_glider_device('microrider', {
                'make': 'Rockland',
                'model': self.excel_metadata['microrider_type'],
                'serial': f'{self.excel_metadata['microrider_sn']}',
                'factory_calibrated':
                self.excel_metadata['microrider_cal'].strftime('%Y-%m-%d'),
                'calibration_date':
                self.excel_metadata['microrider_cal'].strftime('%Y-%m-%d'),
                'calibration_report': '" "',
                'comment': '" "'
            })

        # add from user input (overwrite as necessary)
        if glider_devices is not None:
            for name, value in glider_devices.items():
                self._add_glider_device(name, value)

    def add_netcdf_variables(
            self,
            netcdf_variables: dict[str, dict[str, Any]] = None
    ) -> None:
        """Add NetCDF variable translations and metadata.

        Parameters
        ----------
        netcdf_variables : dict[str, dict[str, Any]], optional
            NetCDF variable name mapping and metadata to overwrite default
            in form {'variable_name': {'attribute': value}

        """
        _log.info('Adding variable metadata to deployment YAML')

        # make sure already have Excel sheet metadata
        self._check_for_excel_meta()

        # initialize common variables
        self.netcdf_variables = {
            'time': {
                'source': 'sci_m_present_time',
                'long_name': 'Time',
                'standard_name': 'time',
                'calendar': 'gregorian',
                'units': 'seconds since 1970-01-01T00:00:00Z',
                'observation_type': 'measured'
            },
            'latitude': {
                'source': 'm_gps_lat',
                'long_name': 'Latitude',
                'standard_name': 'latitude',
                'units': 'degrees_north',
                'comment': 'Estimated between surface fixes',
                'observation_type': 'measured',
                'platform': 'platform',
                'reference': 'WGS84',
                'valid_max': 90.0,
                'valid_min': -90.0,
                '_FillValue': -999.0,
                'coordinate_reference_frame': 'urn:ogc:crs:EPSG::4326'
            },
            'longitude': {
                'source': 'm_gps_lon',
                'long_name': 'Longitude',
                'standard_name': 'longitude',
                'units': 'degrees_east',
                'comment': 'Estimated between surface fixes',
                'observation_type': 'measured',
                'platform': 'platform',
                'reference': 'WGS84',
                'valid_max': 180.0,
                'valid_min': -180.0,
                '_FillValue': -999.0,
                'coordinate_reference_frame': 'urn:ogc:crs:EPSG::4326'
            },
            'heading': {
                'source': 'm_heading',
                'long_name': 'Glider Heading Angle',
                'standard_name': 'platform_orientation',
                'units': 'rad',
                '_FillValue': -999.0
            },
            'pitch': {
                'source': 'm_pitch',
                'long_name': 'Glider Pitch Angle',
                'standard_name': 'platform_pitch_angle',
                'units': 'rad',
                '_FillValue': -999.0
            },
            'roll': {
                'source': 'm_roll',
                'long_name': 'Glider Roll Angle',
                'standard_name': 'platform_roll_angle',
                'units': 'rad',
                '_FillValue': -999.0
            },
            'conductivity': {
                'source': 'sci_water_cond',
                'long_name': 'Conductivity',
                'standard_name': 'sea_water_electrical_conductivity',
                'units': 'S m-1',
                'instrument': 'instrument_ctd',
                'valid_min': 0.0,
                'valid_max': 10.0,
                '_FillValue': -999.0,
                'observation_type': 'measured',
                'accuracy': 0.0003,
                'precision': 0.0001,
                'resolution': 0.00002
            },
            'temperature': {
                'source': 'sci_water_temp',
                'long_name': 'Temperature',
                'standard_name': 'sea_water_temperature',
                'units': 'Celsius',
                'instrument': 'instrument_ctd',
                'valid_min': -5.0,
                'valid_max': 50.0,
                '_FillValue': -999.0,
                'observation_type': 'measured',
                'accuracy': 0.002,
                'precision': 0.001,
                'resolution': 0.0002
            },
            'pressure': {
                'source': 'sci_water_pressure',
                'long_name': 'Pressure',
                'standard_name': 'sea_water_pressure',
                'units': 'dbar',
                'conversion': 'bar2dbar',
                'valid_min': 0.0,
                'valid_max': 2000.0,
                '_FillValue': -999.0,
                'positive': 'down',
                'reference_datum': 'sea-surface',
                'instrument': 'instrument_ctd',
                'observation_type': 'measured',
                'accuracy': 1.0,
                'precision': 2.0,
                'resolution': 0.02,
                'comment': 'ctd pressure sensor'
            },
            'water_velocity_eastward': {
                'source': 'm_water_vx',
                'long_name': 'Depth-Averaged Eastward Sea Water Velocity',
                'standard_name': 'barotropic_eastward_sea_water_velocity',
                'units': 'm s-1',
                '_FillValue': -999.0
            },
            'water_velocity_northward': {
                'source': 'm_water_vy',
                'long_name': 'Depth-Averaged Northward Sea Water Velocity',
                'standard_name': 'barotropic_northward_sea_water_velocity',
                'units': 'm s-1',
                '_FillValue': -999.0
            }
        }

        # add based on devices present in Excel worksheet metadata
        if self.excel_metadata['wetlabs_installed']:
            self._add_netcdf_variable('chlorophyll', {
                'source': 'sci_flbbcd_chlor_units',
                'long_name': 'Chlorophyll',
                'standard_name': 'concentration_of_chlorophyll_in_sea_water',
                'units': 'mg m-3',
                'valid_min': 0.0,
                'valid_max': 50.0,
                '_FillValue': -999.0,
                'resolution': 0.007
            })
            self._add_netcdf_variable('cdom', {
                'source': 'sci_flbbcd_cdom_units',
                'long_name': 'Colored Dissolved Organic Matter',
                'units': 'ppb',
                'valid_min': 0.0,
                'valid_max': 375.0,
                '_FillValue': -999.0,
                'resolution': 0.08
            })
            self._add_netcdf_variable('backscatter_700', {
                'source': 'sci_flbbcd_bb_units',
                'long_name': '700 nm Wavelength Backscatter',
                'units': "1",
                'valid_min': 0.0,
                'valid_max': 5.0,
                '_FillValue': -999.0,
                'resolution': 0.000002
            })
        if self.excel_metadata['oxy_installed']:
            self._add_netcdf_variable('oxygen_concentration', {
                'source': 'sci_oxy4_oxygen',
                'long_name': 'Oxygen Concentration',
                'standard_name': 'mole_concentration_of_dissolved_' +
                                 'molecular_oxygen_in_sea_water',
                'units': 'umol l-1',
                'valid_min': 0.0,
                'valid_max': 500.0,
                '_FillValue': -999.0,
                'accuracy': 8.0,
                'resolution': 1.0
            })
        if self.excel_metadata['par_installed']:
            self._add_netcdf_variable('par', {
                'source': 'sci_bsipar_par',
                'long_name': 'Photosynthetically Active Radiation',
                'standard_name': 'downwelling_photosynthetic_photon_' +
                                 'spherical_irradiance_in_sea_water',
                'units': 'umol m-2 s-1',
                'valid_min': 0.0,
                'valid_max': 6000.0,
                '_FillValue': -999.0
            })
        if self.excel_metadata['bb3_installed']:
            self._add_netcdf_variable('backscatter_470', {
                'source': 'sci_bb3slo_b470_scaled',
                'long_name': '470 nm Wavelength Backscatter',
                'units': "1",
                'valid_min': 0.0,
                'valid_max': 5.0,
                '_FillValue': -999.0,
                'resolution': 0.00001
            })
            self._add_netcdf_variable('backscatter_532', {
                'source': 'sci_bb3slo_b532_scaled',
                'long_name': '532 nm Wavelength Backscatter',
                'units': "1",
                'valid_min': 0.0,
                'valid_max': 5.0,
                '_FillValue': -999.0,
                'resolution': 0.000006
            })
            self._add_netcdf_variable('backscatter_660', {
                'source': 'sci_bb3slo_b660_scaled',
                'long_name': '660 nm Wavelength Backscatter',
                'units': "1",
                'valid_min': 0.0,
                'valid_max': 5.0,
                '_FillValue': -999.0,
                'resolution': 0.0000035
            })
        if self.excel_metadata['lisst_installed']:
            self._add_netcdf_variable('total_volume_concentration', {
                'source': 'sci_lisst_totvol',
                'long_name': 'Total Volume Concentration of Particles',
                'units': 'uL L-1',
                'valid_min': 0.5,
                'valid_max': 700,
                '_FillValue': -999.0,
                'resolution': 0.1
            })
            self._add_netcdf_variable('mean_size', {
                'source': 'sci_lisst_meansize',
                'long_name': 'Mean Particle Size',
                'units': 'um',
                'valid_min': 1.0,
                'valid_max': 500,
                '_FillValue': -999.0
            })
            self._add_netcdf_variable('beam_attenuation', {
                'source': 'sci_lisst_beamc',
                'long_name': 'Beam Attenuation',
                'units': 'm-1',
                'valid_min': 0.3,
                'valid_max': 0.99,
                '_FillValue': -999.0,
                'resolution': 0.1
            })

        # add from user input (overwrite as necessary)
        if netcdf_variables is not None:
            for name, value in netcdf_variables.items():
                self._add_netcdf_variable(name, value)

    def add_profile_variables(
            self,
            profile_variables: dict[str, dict[str, Any]] = None
    ):
        """Add profile-averaged variable metadata

        Parameters
        ----------
        profile_variables : dict[str, dict[str, Any]], optional
            Profile-averaged variable metadata to overwrite default
            in form {'variable_name': {'attribute': value}}

        """
        _log.info('Adding profile variable metadata to deployment YAML')

        # make sure already have Excel sheet metadata
        self._check_for_excel_meta()

        # initialize
        self.profile_variables = {
            'profile_id': {
                'comment': 'Sequential profile number within the ' +
                           'trajectory.  This value is unique in ' +
                           'each file that is part of a single ' +
                           'trajectory/deployment.',
                'long_name': 'Profile ID',
                'valid_max': 2147483647,
                'valid_min': 1,
                '_FillValue': -999.0
            },
            'profile_time': {
                'comment': 'Timestamp corresponding to the mid-' +
                           'point of the profile',
                'long_name': 'Profile Center Time',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'time',
                '_FillValue': -999.0
            },
            'profile_time_start': {
                'comment': 'Timestamp corresponding to the start ' +
                           'of the profile',
                'long_name': 'Profile Start Time',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'time',
                '_FillValue': -999.0
            },
            'profile_time_end': {
                'comment': 'Timestamp corresponding to the end of ' +
                           'the profile',
                'long_name': 'Profile End Time',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'time',
                '_FillValue': -999.0
            },
            'profile_lat': {
                'comment': 'Value is interpolated to provide an ' +
                           'estimate of the latitude at the ' +
                           'mid-point of the profile',
                'long_name': 'Profile Center Latitude',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'latitude',
                'units': 'degrees_north',
                'valid_max': 90.0,
                'valid_min': -90.0,
                '_FillValue': -999.0
            },
            'profile_lon': {
                'comment': 'Value is interpolated to provide an ' +
                           'estimate of the longitude at the ' +
                           'mid-point of the profile',
                'long_name': 'Profile Center Longitude',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'longitude',
                'units': 'degrees_east',
                'valid_max': 180.0,
                'valid_min': -180.0,
                '_FillValue': -999.0
            },
            'u': {
                'comment': 'The depth-averaged current is an ' +
                           'estimate of the net current measured ' +
                           'while the glider is underwater. The ' +
                           'value is calculated over the entire ' +
                           'underwater segment, which may ' +
                           'consist of 1 or more dives.',
                'long_name': 'Depth-Averaged Eastward Sea Water Velocity',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'eastward_sea_water_velocity',
                'units': 'm s-1',
                'valid_max': 10.0,
                'valid_min': -10.0,
                '_FillValue': -999.0
            },
            'v': {
                'comment': 'The depth-averaged current is an ' +
                           'estimate of the net current measured ' +
                           'while the glider is underwater. The ' +
                           'value is calculated over the entire ' +
                           'underwater segment, which may ' +
                           'consist of 1 or more dives.',
                'long_name': 'Depth-Averaged Northward Sea Water Velocity',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'northward_sea_water_velocity',
                'units': 'm s-1',
                'valid_max': 10.0,
                'valid_min': -10.0,
                '_FillValue': -999.0
            },
            'lon_uv': {
                'comment': 'The depth-averaged current is an ' +
                           'estimate of the net current measured ' +
                           'while the glider is underwater. The ' +
                           'value is calculated over the entire ' +
                           'underwater segment, which may ' +
                           'consist of 1 or more dives.',
                'long_name': 'Depth-Averaged Longitude',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'longitude',
                'units': 'degrees_east',
                'valid_max': 180.0,
                'valid_min': -180.0,
                '_FillValue': -999.0
            },
            'lat_uv': {
                'comment': 'The depth-averaged current is an ' +
                           'estimate of the net current measured ' +
                           'while the glider is underwater.  The ' +
                           'value is calculated over the entire ' +
                           'underwater segment, which may ' +
                           'consist of 1 or more dives.',
                'long_name': 'Depth-Averaged Latitude',
                'observation_type': 'calculated',
                'platform': 'platform',
                'standard_name': 'latitude',
                'units': 'degrees_north',
                'valid_max': 90.0,
                'valid_min': -90.0,
                '_FillValue': -999.0
            },
            'time_uv': {
                'comment': 'The depth-averaged current is an ' +
                           'estimate of the net current measured ' +
                           'while the glider is underwater.  The ' +
                           'value is calculated over the entire ' +
                           'underwater segment, which may ' +
                           'consist of 1 or more dives.',
                'long_name': 'Depth-Averaged Time',
                'standard_name': 'time',
                'calendar': 'gregorian',
                'units': 'seconds since 1970-01-01T00:00:00Z',
                'observation_type': 'calculated',
                '_FillValue': -999.0
            },
            'instrument_ctd': {
                'comment': 'pumped CTD',
                'calibration_date':
                self.excel_metadata['ctd_cal'].strftime('%Y-%m-%dT%H:%M:%SZ'),
                'calibration_report': '" "',
                'factory_calibrated':
                self.excel_metadata['ctd_cal'].strftime('%Y-%m-%dT%H:%M:%SZ'),
                'long_name': 'Seabird Glider Payload CTD',
                'make_model': 'Seabird ' +
                self.excel_metadata['ctd_type'],
                'platform': 'platform',
                'serial_number': f'{self.excel_metadata['ctd_sn']}',
                'type': 'platform',
                '_FillValue': -999.0
            },
        }

        # add from user input (overwrite as necessary)
        if profile_variables is not None:
            for name, value in profile_variables.items():
                self._add_profile_variable(name, value)

    def add_qartod_tests(
            self,
            qartod_tests:
            dict[str, dict[str, dict[str, Any]]] = None
    ) -> None:
        """Add QARTOD test parameters

        Note
        ----
        Gross Range Tests use NetCDF variable valid_min, valid_max.

        Spike, Rate of Change, and Flat Line Tests use
        NetCDF variable resolution.

        Parameters
        ----------
        qartod_tests : dict[str, dict[str, dict[str, Any]]], optional
            QARTOD test parameters to overwrite default in form
            {'variable_name': {'test_name': {'parameter_name': value}}}

        """
        _log.info('Adding QARTOD test metadata to deployment YAML')

        # if don't have NetCDF variable metadata already
        if not hasattr(self, 'netcdf_variables'):
            raise AttributeError(
                'Must have NetCDF variable metadata to construct YAML. ' +
                'See method add_netcdf_variables'
            )

        # initialize with instrument limits
        for variable in self.netcdf_variables:
            if (('valid_min' in self.netcdf_variables[variable]) and
               ('valid_max' in self.netcdf_variables[variable])):
                self._add_qartod_test(
                    variable=variable,
                    test='gross_range_test',
                    parameters={'fail_span':
                                [self.netcdf_variables[variable]['valid_min'],
                                 self.netcdf_variables[variable]['valid_max']]
                                })
                if 'resolution' in self.netcdf_variables[variable]:
                    res = self.netcdf_variables[variable]['resolution']
                    self._add_qartod_test(
                        variable=variable,
                        test='spike_test',
                        parameters={'suspect_threshold':
                                    res * 100.0,
                                    'fail_threshold':
                                    res * 200.0
                                    })
                    self._add_qartod_test(
                        variable=variable,
                        test='rate_of_change_test',
                        parameters={'threshold': res * 100.0
                                    })
                    self._add_qartod_test(
                        variable=variable,
                        test='flat_line_test',
                        parameters={'suspect_threshold': 150.0,
                                    'fail_threshold': 300.0,
                                    'tolerance': res * 2.0
                                    })

        # add from user input (overwrite as necessary)
        if qartod_tests is not None and type(qartod_tests) is not bool:
            for variable, test_parameters in qartod_tests.items():
                for test, parameters in test_parameters.items():
                    self._add_qartod_test(variable, test, parameters)

    def construct_yaml(
            self,
            excelsheet: str = None,
            metadata: dict[str, dict[str, Any]] = None,
            glider_devices: dict[str, dict[str, Any]] = None,
            netcdf_variables: dict[str, dict[str, Any]] = None,
            profile_variables: dict[str, dict[str, Any]] = None,
            qartod_tests: dict[str, dict[str, dict[str, Any]]] | bool = None
    ) -> None:
        """Use Excel sheet output to build YAML output.

        Note
        ----
        Method is intended to replace calling other methods individually: if
        other methods called previously, inputs here will overwrite the results

        Paramters
        ---------
        excelsheet : str, optional
            Path to Excel sheet with metadata
        metadata : dict[str, dict[str, Any]], optional
            Global attributes to overwrite default
            in form {'attribute':value}
        glider_devices : dict[str, dict[str, Any]], optional
            Glider device metadata to overwrite default
            in form {'device_name': {'attribute': value}}
        netcdf_variables : dict[str, dict[str, Any]], optional
            NetCDF variable name mapping and metadata to overwrite default
            in form {'variable_name': {'attribute': value}
        profile_variables : dict[str, dict[str, Any]], optional
            Profile-averaged variable metadata to overwrite default
            in form {'variable_name': {'attribute': value}}
        qartod_tests : dict[str, dict[str, dict[str, Any]]] | bool, optional
            QARTOD test parameters to overwrite default in form
            {'variable_name': {'test_name': {'parameter_name': value}}}
            or specify False to skip adding QARTOD tests

        """
        # if gave inputs, overwrite any previous/assign
        if excelsheet is not None:
            self.excelsheet = excelsheet
            if hasattr(self, 'excel_metadata'):
                delattr(self, 'excel_metadata')
        if qartod_tests is not None:
            if hasattr(self, 'qartod_tests'):
                delattr(self, 'qartod_tests')
        else:
            qartod_tests = True
        if metadata is not None:
            if hasattr(self, 'metadata'):
                delattr(self, 'metadata')
        if glider_devices is not None:
            if hasattr(self, 'glider_devices'):
                delattr(self, 'glider_devices')
        if netcdf_variables is not None:
            if hasattr(self, 'netcdf_variables'):
                delattr(self, 'netcdf_variables')
        if profile_variables is not None:
            if hasattr(self, 'profile_variables'):
                delattr(self, 'profile_variables')

        _log.info(f'Creating deployment YAML for {self.ID}')

        # if don't have Excel sheet metadata already
        if not hasattr(self, 'excel_metadata'):
            # if given a sheet to read
            if self.excelsheet is not None:
                # get metadata
                self.add_excel_metadata(self.excelsheet, self.ID)
            # no sheet to read
            else:
                # stop running
                raise AttributeError('Must have Excel worksheet metadata to ' +
                                     'construct YAML. See method collect_' +
                                     'excelsheet_metadata or provide key ' +
                                     '"excelsheet"')

        # if don't already have "metadata": global variables for NetCDF
        if not hasattr(self, 'metadata'):
            self.add_metadata(metadata)

        # if don't have "glider_devices": metadata for installed instruments
        # already
        if not hasattr(self, 'glider_devices'):
            self.add_glider_devices(glider_devices)

        # if don't have "netcdf_variables": metadata for translating glider
        # variables to NetCDF variables already
        if not hasattr(self, 'netcdf_variables'):
            self.add_netcdf_variables(netcdf_variables)

        # if don't have "profile_variables": metadata for profile-averaged
        # variables already
        if not hasattr(self, 'profile_variables'):
            self.add_profile_variables(profile_variables)

        # if don't have qartod tests, if desired, already
        if not hasattr(self, 'qartod_tests'):
            if qartod_tests:
                self.add_qartod_tests(qartod_tests)

        # add everything together
        self.yaml = {
            'metadata': self.metadata,
            'glider_devices': self.glider_devices,
            'netcdf_variables': self.netcdf_variables,
            'profile_variables': self.profile_variables
        }
        if hasattr(self, 'qartod_tests'):
            self.yaml['qartod_tests'] = self.qartod_tests

    def write_yaml(self, outname: str = 'deployment_metadata.yml') -> None:
        """Write yaml dictionary to YAML file.

        Parameters
        ----------
        outname : str, optional
            Full path to YAML file to write

        """
        # assign input
        self.outname = outname

        # make sure have necessary variables
        if not hasattr(self, 'yaml'):
            raise AttributeError('Must have constructed an output yaml ' +
                                 'dictionary. See method `construct_yaml`')

        # write out
        _log.info(f'Wrting deployment YAML for {self.ID} as {self.outname}')
        with open(self.outname, 'w') as outfile:
            yaml.dump(self.yaml, outfile, default_flow_style=False)

    def _check_for_excel_meta(self):
        """Check for Excel metadata already loaded.

        Note
        ----
        Internal; used by many steps

        """
        _log.debug('Checking for existing Excel worksheet metadata')
        # if don't have Excel sheet metadata already
        if not hasattr(self, 'excel_metadata'):
            raise AttributeError('Must have Excel worksheet metadata ' +
                                 'to construct YAML. See method ' +
                                 'collect_excelsheet_metadata')

    def _add_meta(self, name: str, value: Any):
        """Add to dictionary of metadata.

        Note
        ----
        Internal; `metadata` attribute fills in public methods

        Parameters
        ----------
        name : str
            Attribute name
        value : Any
            Attribute value

        """
        _log.debug(f'Adding {name} to metadata')
        self.metadata[name] = value

    def _add_glider_device(self, name: str, value: Any):
        """Add to dictionary of device metadata.

        Note
        ----
        Internal; `glider_devices` attribute fills in public methods

        Parameters
        ----------
        name : str
            Attribute name
        value : Any
            Attribute value

        """
        _log.debug(f'Adding {name} to device metadata')
        self.glider_devices[name] = value

    def _add_netcdf_variable(self, name: str, value: Any):
        """Add to dictionary of variable mapping metadata.

        Note
        ----
        Internal; `netcdf_variables` attribute fills in public methods

        Parameters
        ----------
        name : str
            Attribute name
        value : Any
            Attribute value

        """
        _log.debug(f'Adding {name} to variable mapping metadata')
        self.netcdf_variables[name] = value

    def _add_profile_variable(self, name: str, value: Any):
        """Add to dictionary of profile variable mapping metadata.

        Note
        ----
        Internal; `profile_variables` attribute fills in public methods

        Parameters
        ----------
        name : str
            Attribute name
        value : Any
            Attribute value

        """
        _log.debug(f'Adding {name} to profile variable mapping metadata')
        self.profile_variables[name] = value

    def _add_qartod_test(self, variable: str, test: str, parameters: Any):
        """Add to dictionary of QARTOD tests to perform.

        Note
        ----
        Internal; `qartod_tests` attribute fills in public methods

        Parameters
        ----------
        variable : str
            Variable name
        test : str
            Test name for variable
        parameters : Any
            Test parameters for variable

        """
        _log.debug(f'Adding {variable} to QARTOD test metadata')
        try:
            self.qartod_tests['streams'][variable]['qartod'][test] = parameters
        except Exception:
            try:
                self.qartod_tests['streams'][variable] = {
                    'qartod': {test: parameters}
                }
            except Exception:
                self.qartod_tests = {'streams': {variable: {
                    'qartod': {test: parameters}
                }}}
        # #this part will only be necessary/useful if pyglider and
        # ioos_qc cooperate
        # try:
        #     self.netcdf_variables[variable]['ancillary_variables'] +=
        # f'{variable}_qartod_{test} '
        # except:
        #     self.netcdf_variables[variable]['ancillary_variables'] =
        # f' {variable}_qartod_{test} '


class DeploymentNetCDF():
    """Contains deployment-specific data and metadata in IOOS Glider DAC form

    Parameters
    ----------
    main_directory : str
        Path to directory with raw file directory and cache file
        directory that will also hold the processed file directory
    binary_directory : str, default = 'Raw'
        Path, relative to main_directory, that holds raw files
    cache_directory : str, default = join_path('Raw', 'Cache')
        Path, relative to main_directory, that holds cache files
    deployment_yaml : str, default = 'deployment_metadata.yml'
        Path, relative to main_directory, for the YAML with the
        deployment-specific metadata
    style : str, default = 'Realtime'
        Deployment processing type, either 'Realtime' or 'Delayed'

    Attributes
    ----------
    main_directory : str
        Full path to raw, cache, and processed directories
    binary_directory : str
        Full path to directory with raw files
    cache_directory : str
        Full path to directory with cache files
    deployment_yaml : str
        Full path to deployment-specific YAML file
    style : str
        Deployment processing type
    l0timeseries_directory : str
        Full path to the directory with the L0 timeseries NetCDFs
    l0profile_directory : str
        Full path to the directory with the L0 profile NetCDFs
    l0timeseries_outname : str
        Full path to the L0 timeseries NetCDF
    l1timeseries_directory : str
        Full path to the directory with the L1 timeseries NetCDFs
    l1profile_directory : str
        Full path to the directory with the L1 profile NetCDFs
    l1timeseries_outname : str
        Full path to the L1 timeseries NetCDF
    passing_state : dict
        Full path of each file checked and its passing state,
        which is False if not passed or already evaluated

    """

    def __init__(
            self,
            main_directory: str,
            binary_directory: str = 'Raw',
            cache_directory: str = join_path('Raw', 'Cache'),
            deployment_yaml: str = 'deployment_metadata.yml',
            style: str = 'Realtime'
    ) -> None:
        self.main_directory = main_directory
        self.binary_directory = join_path(main_directory, binary_directory)
        self.cache_directory = join_path(main_directory, cache_directory)
        self.deployment_yaml = join_path(
            main_directory, style, deployment_yaml
        )
        self.style = style

    def read_timeseries(self, timeseries_file: str) -> xr.Dataset:
        """Read a timeseries NetCDF as a Dataset

        Parameters
        ----------
        timeseries_file : str
            Path, relative to the class's main_directory and style,
            to a timeseries NetCDF file

        Returns
        -------
        xarray Dataset
            Dataset from file

        """
        timeseries_file = join_path(
            self.main_directory, self.style, timeseries_file
        )
        _log.info(f'Reading {timeseries_file} into memory')
        with xr.open_dataset(timeseries_file) as ds:
            return ds

    def read_profiles(self, profile_directory: str) -> xr.Dataset:
        """Read a directory of profile NetCDFs as a single Dataset

        Parameters
        ----------
        profile_directory : str
            Path, relative to the class's main_directory and style,
            to a directory of profile NetCDF files

        Returns
        -------
        xarray Dataset
            Combined Dataset from files

        """
        profile_directory = join_path(
            self.main_directory, self.style, profile_directory, '*.nc'
        )
        _log.info(f'Reading all files in {profile_directory} into memory')
        with xr.open_mfdataset(profile_directory) as ds:
            return ds

    def make_L0(
            self,
            l0timeseries_directory: str = 'L0-timeseries',
            l0profile_directory: str | bool | None = 'L0-profiles'
    ) -> None:
        """Create L0 (read and interpolated) timeseries and profile NetCDFs.

        Note
        ----
        Reading and interpolating occur according to `pyglider`

        Parameters
        ----------
        l0timeseries_directory : str, optional
            Path, relative to class's main_directory and style,
            to the directory to put the L0 timeseries NetCDFs
        l0profile_directory : str, optional
            Path, relative to class's main_directory and style,
            to the directory to put the L0 profile NetCDFs.
            Specify None or False to skip writing profile NetCDFs

        """
        # assign inputs
        match self.style:
            case 'Realtime':
                self.search = '*.[s|t]bd'
                self.profile_filt_time = 20
                self.profile_min_time = 60
            case 'Delayed':
                self.search = '*.[d|e]bd'
                self.profile_filt_time = 100
                self.profile_min_time = 300
            case _:
                raise ValueError('Input "style" must be "realtime", ' +
                                 '"delayed", or None')
        self.l0timeseries_directory = join_path(
            self.main_directory, self.style, l0timeseries_directory
        )
        if l0profile_directory:
            self.l0profile_directory = join_path(
                self.main_directory, self.style, l0profile_directory
            )
        else:
            self.l0profile_directory = l0profile_directory

        # turn binary *.*bd (file extension based on search) into a single
        # timeseries netcdf file
        self.l0timeseries_outname = slocum.binary_to_timeseries(
            self.binary_directory, self.cache_directory,
            self.l0timeseries_directory, self.deployment_yaml,
            search=self.search, profile_filt_time=self.profile_filt_time,
            profile_min_time=self.profile_min_time
        )
        _log.info('Created L0 single timeseries NetCDF' +
                  f' {self.l0timeseries_outname}')

        # make profile netcdf files
        if self.l0profile_directory:
            _log.info('Creating L0 profile NetCDFs in ' +
                      f'{self.l0profile_directory}')
            extract_timeseries_profiles(
                self.l0timeseries_outname,
                self.l0profile_directory,
                self.deployment_yaml
            )

    def make_L1(
            self,
            l1timeseries_directory: str = 'L1-timeseries',
            l1profile_directory: str | bool | None = 'L1-profiles'
    ) -> None:
        """Create the L1 (QARTOD tested) timeseries and profile NetCDFs.

        Note
        ----
        Tests performed according to `ioos_qc`

        Parameters
        ----------
        l1timeseries_directory : str, optional
            Path, relative to class's main_directory and style,
            to the directory to put the L1 timeseries NetCDFs
        l1profile_directory : str, optional
            Path, relative to class's main_directory and style,
            to the directory to put the L1 profile NetCDFs.
            Specify None or False to skip writing profile NetCDFs

        """
        # make sure already have an L0
        if not hasattr(self, 'l0timeseries_outname'):
            raise AttributeError(
                'No L0 timeseries file path exists. ' +
                'Must run make_L0 before make_L1'
            )

        # assign inputs
        self.l1timeseries_directory = join_path(
            self.main_directory, self.style, l1timeseries_directory
        )
        if l1profile_directory:
            self.l1profile_directory = join_path(
                self.main_directory, self.style, l1profile_directory
            )
        else:
            self.l1profile_directory = l1profile_directory

        # make sure output directory exists
        if not exists(self.l1timeseries_directory):
            makedirs(self.l1timeseries_directory)

        # create output name based on L0 name
        self.l1timeseries_outname = join_path(
            self.l1timeseries_directory,
            basename(self.l0timeseries_outname)
        )
        _log.info('Creating L1 single timeseries ' +
                  f'NetCDF {self.l1timeseries_outname}')

        # use yaml configuration for ioos_qc configuration
        _log.info('Extracting QARTOD test parameters from ' +
                  f'{self.deployment_yaml}')
        with open(self.deployment_yaml) as fin:
            metadata = yaml.safe_load(fin)
        config = Config(metadata['qartod_tests'])

        # perform qartod tests on L0 to create L1
        _log.info(f'Running QARTOD tests on {self.l0timeseries_outname}')
        with xr.open_dataset(self.l0timeseries_outname) as ds:
            # run initial tests
            stream = XarrayStream(ds, time='time', z='depth',
                                  lat='latitude', lon='longitude')
            runner = list(stream.run(config))
            # group initial results by stream_id (source variable)
            grouped_runner = (
                {k: list(g) for k, g in groupby(runner, lambda r: r.stream_id)}
            )
            # run aggregate tests and add to initial results
            for source, run in grouped_runner.items():
                result = collect_results(run)
                agg = ContextResult(
                    stream_id=source,
                    results=[CallResult(
                        package='',
                        test='qc',
                        function=aggregate,
                        results=aggregate(result)
                    )],
                    subset_indexes=run[0].subset_indexes,
                    data=run[0].data,
                    tinp=run[0].tinp,
                    zinp=run[0].zinp,
                    lat=run[0].lat,
                    lon=run[0].lon
                )
                runner.append(agg)
            # store results
            store = PandasStore(runner)
            # append results to existing Dataset
            ds = xr.merge([
                ds,
                # store results as Pandas DataFrame
                store.save()
                # set the index (xarray coordinate) to time
                .set_index('time')
                # remove redundant columns added by ioos_qc
                .drop(columns=['z', 'lat', 'lon'])
                # make all bytes as required by GDAC
                .astype('int8')
                # convert results to Dataset
                .to_xarray()
            ])
        # redefine some aggregate results how GDAC wants
        for name, da in ds.data_vars.items():
            if '_qc' in name:
                da[da == 2] = 0
        # add metadata (loosely based on ioos_qc CFNetCDFStore)
        for result in store.collected_results:
            # get variable name of focus test result
            name = column_from_collected_result(result)
            _log.debug(f'Adding {name} to dataset')
            # get test function of focus test result
            func = result.function
            # get name of source variable
            source = result.stream_id
            # get existing ancillary variables for source
            ancillary = getattr(ds[source], 'ancillary_variables', '')
            # add this one
            ancillary += f' {name}'
            # if aggregate test result
            if '_qc' in name:
                # define flags
                flag_values = [x for x in range(0, 10)]
                flag_names = (
                    'no_qc_performed good_data probably_good_data' +
                    'bad_data_that_are_potentially_correctable ' +
                    'bad_data value_changed not_used not_used ' +
                    'interpolated_value missing_value'
                )
                # define "input" for focus test result
                call = 'various'
            else:
                # separate flag details
                flags = getmodule(func).FLAGS
                flag_names = [d for d in flags.__dict__ if not
                              d.startswith("__")]
                flag_values = [getattr(flags, d) for d in flag_names]
                # get qartod input for focus test result
                call = config.calls_by_stream_id(source)
                call = [c for c in call if c.module == result.package
                        and c.method == result.test]
                call = str(call[0].kwargs)
            # update Dataset
            ds[source] = ds[source].assign_attrs(ancillary_variables=ancillary)
            ds[name] = ds[name].assign_attrs(
                standard_name=getattr(func, 'standard_name', 'quality_flag'),
                long_name=getattr(func, 'long_name', 'Quality Flag'),
                flag_values=np.byte(flag_values),
                flag_meanings=" ".join(flag_names),
                valid_min=np.byte(min(flag_values)),
                valid_max=np.byte(max(flag_values)),
                ioos_qc_module=result.package,
                ioos_qc_test=result.test,
                ioos_qc_target=source,
                ioos_qc_config=call,
                _FillValue=np.byte(-127)
            )
        # export L1 timeseries netcdf file
        _log.info(f'Saving QARTOD test results as {self.l1timeseries_outname}')
        ds.to_netcdf(self.l1timeseries_outname)

        # make L1 profile NetCDF files for IOOS GliderDAC
        if self.l1profile_directory:
            _log.info('Creating L1 profile NetCDFs in' +
                      f' {self.l1profile_directory}')
            extract_timeseries_profiles(self.l1timeseries_outname,
                                        self.l1profile_directory,
                                        self.deployment_yaml)

    def check_compliance(self, profile_directory: str = 'L1-profiles') -> None:
        """Use the IOOS compliance checker on profile files.

        Parameters
        ----------
        profile_directory : str, optional
            Path, relative to class's main_directory and style,
            to the directory with the target profile NetCDFs

        """
        profile_directory = join_path(self.main_directory, self.style,
                                      profile_directory)
        _log.info(f'Checking profile NetCDFs in {profile_directory}')

        # define log level for checker verbosity
        match _log.root.level:
            case 10:
                # DEBUG
                log_level = 2
            case 20:
                # INFO
                log_level = 1
            case _:
                # everything else
                log_level = 0

        # get file names to check
        file_names = listdir(profile_directory)
        file_names = [file for file in file_names if '.nc' in file]

        # initialize compliance checker
        check_suite = CheckSuite()
        check_suite.load_all_available_checkers()
        checker = ComplianceChecker()

        # run compliance checker for each file
        passed = [False] * len(file_names)
        for idx, file in enumerate(file_names):
            file = file.split('.')[0]
            file = join_path(profile_directory, file)
            _log.info(f'Checking {file}_report.txt')
            if not exists(file+'_report.txt'):
                passed[idx], _ = checker.run_checker(
                    ds_loc=file+'.nc',
                    output_filename=file+'_report.txt',
                    checker_names=['gliderdac'],
                    verbose=log_level,
                    criteria='normal'
                )
                if passed[idx]:
                    isnot = 'is'
                else:
                    isnot = 'is not'
                _log.info(f'{file}.nc {isnot} IOOS GliderDAC compliant.' +
                          f'See {file}_report.txt for details.')

        # output
        self.passing_state = {
            join_path(profile_directory, file):
            state for file, state in zip(file_names, passed)
        }

    def create_summary(
            self,
            timeseries_file: str = 'L0',
            output_file: str = 'L0',
            display: bool = False,
            author: str = 'Anonymous',
            extra_text: str = '',
            map_bounds: list[float] = None,
            globe_position: str = 'BL',
            plots: tuple[dict[str, str]] = (
                {'source': 'temperature',
                 'cmap': 'cmocean.sequential.Thermal_20'},
                {'source': 'salinity',
                 'cmap': 'cmocean.sequential.Haline_20'},
                {'source': 'density',
                 'cmap': 'cmocean.sequential.Dense_20'}
            )
    ):
        """Output a summary page from a timeseries NetCDF file.

        Parameters
        ----------
        timeseries_file : str, optional
            Path, relative to class's main_directory and style,
            to target timeseries NetCDF file.
            May use shortcuts 'L0' and 'L1' to refer to files
            created with class methods.
        output_file : str, optional
            Path, relative to class's main_directory and style,
            for output PNG file.
            May use shortcuts 'L0' and 'L1' to refer to files
            created with class methods.
            If False, will not write a file.
        display : bool, optional
            Whether to display the summary in default viewer
        author : str, optional
            Name of the person creating the summary page
        extra_text : str, optional
            Any text to add on the same line as funding acknowledgement
            (ie glider faults)
        map_bounds : list or None, optional
            Map bounds in form
            [minimum latitude, maximum latitude,
            minimum longitude, maximum longitude].
            Specify None to define based on data
        globe_position : str, optional
            Position on map for overview globe. Use codes in terms of
            (T)op, (M)iddle, (B)ottom and
            (L)eft, (C)enter, (R)ight
        plots : tuple[dict[str, str]], optional
            Three subplots, in order of top to bottom,
            with palettable colortables in form
            {'source': 'variable_name',
            'cmap': 'path.to.colortable'}

        """
        # re-define inputs
        match timeseries_file:
            case 'L0':
                if not hasattr(self, 'l0timeseries_outname'):
                    raise AttributeError(
                        'DeploymentNetCDF class does not have attribute ' +
                        '"l0timeseries_outname". Must run ' +
                        'make_l0 method to use ' +
                        'timeseries_file input "L0".'
                    )
                timeseries_file = self.l0timeseries_outname
            case 'L1':
                if not hasattr(self, 'l1timeseries_outname'):
                    raise AttributeError(
                        'DeploymentNetCDF class does not have attribute ' +
                        '"l1timeseries_outname". Must run ' +
                        'make_l1 method to use ' +
                        'timeseries_file input "L1".'
                    )
                timeseries_file = self.l1timeseries_outname
            case _:
                timeseries_file = join_path(
                    self.main_directory, self.style,
                    timeseries_file
                )
        match output_file:
            case 'L0':
                if not hasattr(self, 'l0timeseries_outname'):
                    raise AttributeError(
                        'DeploymentNetCDF class does not have attribute ' +
                        '"l0timeseries_outname". Must run ' +
                        'make_l0 method to use ' +
                        'output_file input "L0".'
                    )
                output_file = join_path(
                    self.main_directory, self.style,
                    Path(self.l0timeseries_outname).stem + '.png'
                )
            case 'L1':
                if not hasattr(self, 'l1timeseries_outname'):
                    raise AttributeError(
                        'DeploymentNetCDF class does not have attribute ' +
                        '"l1timeseries_outname". Must run ' +
                        'make_l1 method to use ' +
                        'output_file input "L1".'
                    )
                output_file = join_path(
                    self.main_directory, self.style,
                    Path(self.l1timeseries_outname).stem + '.png'
                )
            case False:
                pass
            case _:
                output_file = join_path(
                    self.main_directory, self.style,
                    output_file
                )

        _log.info('Creating summary page')

        # get data and metadata
        data = self.read_timeseries(timeseries_file)
        with open(self.deployment_yaml) as fin:
            metadata = yaml.safe_load(fin)
            metadata = metadata['metadata']

        # create additional header info
        title = ('Ocean Glider Deployment Summary: ' +
                 f'{metadata['deployment_name']} - {metadata['project']}')
        date = datetime.today().strftime('%d %B, %Y')
        extra_text = f'{metadata['acknowledgement']}. {extra_text}'

        # #create "snapshot": text table of select metadata
        # start/end location
        loc = []
        lat = data['latitude'].values
        lon = data['longitude'].values
        for lat, lon in zip([first_nonnan(lat), last_nonnan(lat)],
                            [first_nonnan(lon), last_nonnan(lon)]):
            lat_deg, lat_min = dd2dm(lat)
            if lat_deg < 0:
                ns = 'S'
                lat_deg *= -1
            else:
                ns = 'N'
            lon_deg, lon_min = dd2dm(lon)
            if lon_deg < 0:
                ew = 'W'
                lon_deg *= -1
            else:
                ew = 'E'
            loc.append(f"{lat_deg:.0f}\xb0{lat_min:.02f}'" +
                       f"{ns},{lon_deg:.0f}\xb0{lon_min:.02f}'{ew}")
        # deployment duration
        duration = data['time'].values[-1]-data['time'].values[0]
        dys = duration.astype('timedelta64[D]').astype(np.int32)
        hrs = duration.astype('timedelta64[h]').astype(np.int32) % 24
        duration = f'{dys} days, {hrs} hours'
        # science sensors output
        sci_vars = list(data.data_vars)
        [sci_vars.remove(v) for v in ['latitude', 'longitude', 'heading',
                                      'pitch', 'roll', 'pressure', 'depth',
                                      'water_velocity_eastward',
                                      'water_velocity_northward',
                                      'distance_over_ground', 'profile_index',
                                      'profile_direction']]
        sci_vars = [data[v].attrs['long_name'] for v in sci_vars]
        # construct
        snapshot = (
            [
                'Deployment date',
                'Deployment location',
                'Retrieval date',
                'Retrieval location',
                'Deployment duration',
                'Science sensors'] +
            [' ' for _ in range(len(sci_vars)-1)] +
            ['Number of profiles',
                'Glider name',
                'Profiling range',
                'Max depth reached',
                'Distance covered'
             ],
            [
                np.datetime_as_string(data['time'].values[0], unit='s',
                                      timezone='UTC'),
                loc[0],
                np.datetime_as_string(data['time'].values[-1], unit='s',
                                      timezone='UTC'),
                loc[1],
                duration] +
            sci_vars + [f'{np.nanmax(data['profile_index'].values):.0f}',
                        metadata['glider_name'],
                        f'0-{metadata['glider_pump']}',
                        f'{np.nanmax(data['depth'].values):.02f}m',
                        f'{data['distance_over_ground'].values[-1]:.02f}km'
                        ]
        )

        # define map limits
        if map_bounds is not None:
            min_lat = map_bounds[0]
            max_lat = map_bounds[1]
            min_lon = map_bounds[2]
            max_lon = map_bounds[3]
        else:
            min_lat = np.floor((np.nanmin(data['latitude'].values) - 0.5)
                               * 10) / 10
            max_lat = np.ceil((np.nanmax(data['latitude'].values) + 0.5)
                              * 10) / 10
            min_lon = np.floor((np.nanmin(data['longitude'].values) - 1)
                               * 10) / 10
            max_lon = np.ceil((np.nanmax(data['longitude'].values) + 1)
                              * 10) / 10

        # initialize summary page
        fig = pygmt.Figure()
        # define gmt processing debug output level based on
        # user's selected log level
        match _log.root.level:
            # DEBUG
            case 10:
                # Debugging
                log_level = 'd'
            # everything else
            case _:
                # Warnings (GMT default)
                log_level = 'w'
        # define gmt setup
        pygmt.config(
            PS_MEDIA='a4', FONT_ANNOT_PRIMARY=8, FONT_LABEL=8,
            PROJ_LENGTH_UNIT='c', MAP_FRAME_TYPE='plain', MAP_TICK_PEN='black',
            MAP_FRAME_PEN='thinnest,black', GMT_VERBOSE=log_level
        )
        # define map common paramters
        avg_lat = np.mean([min_lat, max_lat])
        avg_lon = np.mean([min_lon, max_lon])
        z_levels = ','.join(
            [str(num) for num in [*range(0, 901, 100)] +
             [*range(1000, 6001, 1000)]]
        )
        zone = from_latlon(avg_lat, avg_lon)
        zone = f'{zone[2]}{zone[3]}'
        # define Hovmoller diagram common parameters
        min_time = np.nanmin(data['time'].values)
        max_time = np.nanmax(data['time'].values)
        min_depth = 0
        max_depth = np.nanmax(data['depth'].values)
        region = [min_time, max_time, min_depth, max_depth]
        height = 5.25
        projection = f'X18/-{height}'
        yshifts = [0.5, 6, 11.5]
        xaxis = ['S', 's', 's']
        _log.debug('Using map and plot boundaries\n' +
                   '\tminimum\n' +
                   f'\t\tlatitude {min_lat}\n' +
                   f'\t\tlongitude {min_lon}\n' +
                   f'\t\ttime {min_time}\n' +
                   f'\t\tdepth {min_depth}\n' +
                   '\tmaximum\n' +
                   f'\t\tlatitude {max_lat}\n' +
                   f'\t\tlongitude {max_lon}\n' +
                   f'\t\ttime {max_time}\n' +
                   f'\t\tdepth {max_depth}\n' +
                   '\taverage\n' +
                   f'\t\tlatitude {avg_lat}\n' +
                   f'\t\tlongitude {avg_lon}'
                   )
        # Hovmoller diagrams
        fig.shift_origin(xshift='f1')
        for dy, var, s in zip(yshifts, plots[::-1], xaxis):
            z = data[var['source']].values
            lbl = (f'{data[var['source']].attrs['long_name']} ' +
                   f'({data[var['source']].attrs['units']})')
            _log.info(f'Adding {lbl} plot to summary page')
            pygmt.makecpt(
                cmap=temporary_cpt(palette=var['cmap']),
                background='o',
                series=[np.nanmean(z)-3*np.nanstd(z),
                        np.nanmean(z)+3*np.nanstd(z)]
            )
            fig.shift_origin(yshift=f'f{dy}')
            with pygmt.config(FONT_ANNOT_PRIMARY=6):
                fig.basemap(region=region, projection=projection,
                            frame=['af', 'y+lDepth (m)', 'We+ggray90'])
            fig.basemap(frame=['af', f'{s}n'])
            fig.plot(
                x=data['time'].values,
                y=data['depth'].values,
                fill=z, cmap=True,
                style='c0.5p'
            )
            with pygmt.config(FONT_ANNOT_PRIMARY=10, FONT_LABEL=12):
                fig.colorbar(cmap=True, frame=['af', f'x+l{lbl}'],
                             position=f'jBR+w{height}+o-0.5/0')
        # main map
        _log.info('Adding map to summary page')
        grid = pygmt.datasets.load_earth_relief(
            region=[min_lon-1, max_lon+1, min_lat-1, max_lat+1],
            resolution='15s', data_source='gebcosi'
        ) * -1
        fig.shift_origin(xshift='f0.5', yshift='f17.25')
        pygmt.makecpt(cmap=temporary_cpt(
            palette='colorbrewer.sequential.Blues_9'), background='o',
            series=z_levels)
        fig.grdimage(
            region=f'{min_lon}/{min_lat}/{max_lon}/{max_lat}+r',
            projection=f'U{zone}/10.25+du', frame=['wEsN', 'af'],
            grid=grid, cmap=True
        )
        fig.coast(land='white', shorelines=True)
        pygmt.makecpt(
            cmap=temporary_cpt(palette='cmocean.sequential.Amp_20'),
            background='o',
            series=[min_time, max_time]
        )
        fig.plot(
            x=data['longitude'].values,
            y=data['latitude'].values,
            fill=data['time'].values, cmap=True,
            style='c2p'
        )
        fig.basemap(frame='lrbt')
        # inset map
        grid = pygmt.datasets.load_earth_relief(
            resolution='10m', data_source='gebcosi'
        ) * -1
        with fig.inset(position=f'j{globe_position}+w2.75+o0.25'):
            pygmt.makecpt(
                cmap=temporary_cpt(palette='colorbrewer.sequential.Blues_9'),
                background='o',
                series=z_levels
            )
            fig.grdimage(
                region='g', projection=f'G{avg_lon}/{avg_lat}/10/?',
                frame='g10', grid=grid, cmap=True
            )
            fig.coast(land='white', shorelines=True)
            fig.plot(
                x=[max_lon, max_lon, min_lon],
                y=[min_lat, max_lat, max_lat],
                close=f'+x{min_lon}+y{min_lat}+pthin,red'
            )
        # metadata text box
        _log.info('Adding metadata snapshot to summary page')
        _log.debug(f'Using text\n{snapshot}')
        fig.shift_origin(xshift='f13.75', yshift='f17.25')
        fig.basemap(
            region=[0, 1, 0, len(snapshot[0])+1],
            projection='X7/-10.5', frame='lbtr+ggray95'
        )
        for col, text in enumerate(snapshot):
            fig.text(
                x=np.ones(len(text))*0.375*col, y=range(1, len(text)+1),
                text=text, justify='LM', offset='0.25/0', font=7
            )
        # main map colorbar ("attached" to metadata text box for easier
        # position and sizing)
        pygmt.makecpt(
            cmap=temporary_cpt(palette='cmocean.sequential.Amp_20'),
            background='o', series=[min_time, max_time]
        )
        fig.colorbar(cmap=True, frame='af', position='jBL+w10.5+o-0.6/0+ma')
        # title, author, date, extra text
        _log.info('Adding header information to summary page')
        debug_text = 'Using\n' +\
            f'\ttitle {title}\n' +\
            f'\tauthor {author}\n' +\
            f'\tdate {date}\n' +\
            f'\ttext {extra_text}'
        _log.debug(debug_text)
        fig.shift_origin(xshift='f0', yshift='f28')
        fig.text(
            region=[-1, 1, -1.5, 1.5],
            projection='X21/1.75',
            x=0, y=0.5,
            text=title, font=12, justify='BC'
        )
        fig.text(
            x=0, y=0,
            text=(
                f'Deployed from {(data['time'].values[0]
                                  .astype('datetime64[s]').item()
                                  .strftime('%d %B, %Y'))} ' +
                (f'to {(data['time'].values[-1]
                        .astype('datetime64[s]').item()
                        .strftime('%d %B, %Y'))} ' +
                 f'*** Prepared by {author} on {date}')
            ),
            justify='MC'
        )
        with NamedTemporaryFile(mode='w', suffix='.txt',
                                delete_on_close=False) as f:
            f.write(f'> 0 -0.5 8p 18 c\n{extra_text}')
            f.close()
            fig.text(textfiles=f.name, font=8, justify='TC', M=True)

        # save
        if output_file:
            _log.info(f'Saving summary page as {output_file}')
            fig.savefig(output_file, crop=False)
        # display
        if display:
            _log.info('Displaying summary page')
            fig.show(crop=False)
